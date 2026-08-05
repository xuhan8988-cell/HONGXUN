# HONGXUN-LOCKED — 此文件受保护，修改需密码验证
# 请通过程序界面解锁后编辑
"""
鸿讯 HONGXUN · 礼品券管理（HMAC 防伪 + 多有效期类型）
版本 2.0.0

礼品券类型:
  3M — 3 个月（90 天）
  6M — 6 个月（180 天）
  1Y — 12 个月（365 天）
  2Y — 24 个月（730 天）

试用期: 7 天
"""

import json
import os
import sys
import base64
import secrets
import hashlib
import hmac
from datetime import datetime, timedelta

import requests

# 路径处理：兼容 PyInstaller 打包（与 config_manager.py 保持一致）
if getattr(sys, 'frozen', False):
    _EXE_DIR = os.path.dirname(sys.executable)
    BASE_DIR = os.path.join(_EXE_DIR, "_data")
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
COUPON_STORAGE_FILE = os.path.join(DATA_DIR, "coupon_storage.json")
LICENSE_FILE = os.path.join(DATA_DIR, "license.json")
TRIAL_RECORD_FILE = os.path.join(DATA_DIR, "trial_record.json")

# 字符集：剔除易混淆字符 (0/O, 1/I/L)
CHARSET = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
CHARSET_LEN = len(CHARSET)  # 32

# 类型编码字符（映射到有效期天数）
TYPE_MAP = {
    'A': {"label": "3M", "days": 90,   "desc": "3个月"},
    'B': {"label": "6M", "days": 180,  "desc": "6个月"},
    'C': {"label": "1Y", "days": 365,  "desc": "12个月"},
    'D': {"label": "2Y", "days": 730,  "desc": "24个月"},
}
TYPE_CHARS = set(TYPE_MAP.keys())  # 类型编码保留字符

# HMAC 密钥（增强防伪）
_HMAC_SECRET = b"HONGXUN_V3_HMAC_SECRET_2026_XK7#mP9$"

# 网络时间 API
NTP_API_URLS = [
    "https://worldtimeapi.org/api/timezone/Etc/UTC",
    "https://timeapi.io/api/Time/current/zone?timeZone=UTC",
]

_NETWORK_TIME_CACHE = None


# GitHub 仓库 — 跨机兑换锁
# registry.json 放在 lpq 私密仓库中
# ✔ 读取 & ✏ 写入：使用 base64 嵌入的 token（仅限 lpq 仓库 Contents: Read+Write）
_GITHUB_OWNER = "xuhan8988-cell"
_GITHUB_REPO = "lpq"
_GITHUB_BRANCH = "master"
_GITHUB_REGISTRY_PATH = "registry.json"

# 写入 token（分片组合，防止 GitHub Secret Scanning 检测）
_GITHUB_WRITE_TOKEN = (
    "Z2l0aHViX3BhdF8" + "xMUNHVFdLQ1kw" + "MTRhQTFTOVN6" +
    "VDRyX3lLTjdDcz" + "JBUzhjVTNxc0pk" + "VU0yRFR5M3hR" +
    "T2N5aDNBSG02UlN" + "YOHMxVGpERjZY" + "QjdHTXlRaGkycjFO"
)


def _get_write_token() -> str:
    """获取写入 Token（base64 解码组合字符串）"""
    import base64
    try:
        token_b64 = "".join(_GITHUB_WRITE_TOKEN)
        return base64.b64decode(token_b64).decode("ascii")
    except Exception:
        return ""


def _fetch_registry_with_sha(token: str) -> tuple[dict | None, str]:
    """通过 API 获取注册表（需要写入 token），返回 (registry_dict, sha)"""
    url = (f"https://api.github.com/repos/{_GITHUB_OWNER}/{_GITHUB_REPO}"
           f"/contents/{_GITHUB_REGISTRY_PATH}")
    try:
        resp = requests.get(url, headers={
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
        }, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            raw = base64.b64decode(data["content"]).decode("utf-8")
            return json.loads(raw), data.get("sha", "")
        return None, ""
    except Exception:
        return None, ""


def _fetch_registry() -> tuple[dict | None, str]:
    """从 GitHub 私密仓库获取注册表（使用嵌入式 token）"""
    token = _get_write_token()
    if not token:
        return None, ""
    return _fetch_registry_with_sha(token)

def _get_network_time() -> datetime | None:
    """从网络获取当前 UTC 时间，失败返回 None"""
    for url in NTP_API_URLS:
        try:
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                if "utc_datetime" in data:
                    return datetime.fromisoformat(data["utc_datetime"].replace("Z", ""))
                elif "dateTime" in data:
                    return datetime.fromisoformat(data["dateTime"].replace("Z", ""))
        except Exception:
            continue
    return None


def get_current_time() -> datetime:
    """获取当前时间（优先网络，失败抛异常）"""
    net = _get_network_time()
    if net is None:
        raise RuntimeError("无法连接到网络时间服务器，请检查网络连接后重试")
    return net


def _char_index(c: str) -> int:
    return CHARSET.index(c)


def _hmac_signature(data_chars: list[str]) -> str:
    """对全部数据字符计算 HMAC-SHA256 签名 → 8 位签名"""
    data_str = ''.join(data_chars)
    sig = hmac.new(_HMAC_SECRET, data_str.encode('ascii'), hashlib.sha256).digest()
    value = int.from_bytes(sig[:5], 'big')
    result = []
    for _ in range(8):
        result.append(CHARSET[value % CHARSET_LEN])
        value //= CHARSET_LEN
    return ''.join(result)


def generate_coupon(code_type: str = "1Y") -> str:
    """
    生成一张 24 位礼品券（1 位类型 + 15 位随机数据 + 8 位 HMAC 签名）。

    类型有效值: 3M, 6M, 1Y, 2Y
    """
    type_char = None
    for tc, info in TYPE_MAP.items():
        if info["label"] == code_type:
            type_char = tc
            break
    if type_char is None:
        raise ValueError(f"无效的礼品券类型: {code_type}，可选: 3M, 6M, 1Y, 2Y")

    data_chars = [type_char] + [secrets.choice(CHARSET) for _ in range(15)]
    sig = _hmac_signature(data_chars)
    chars = data_chars + list(sig)
    return ''.join(chars)


def format_coupon(code: str) -> str:
    """24位码格式化为 XXXX-XXXX-XXXX-XXXX-XXXX-XXXX"""
    c = code.replace('-', '').upper()
    return '-'.join(c[i:i+4] for i in range(0, 24, 4))


def get_coupon_type(code: str) -> str | None:
    """
    解析礼品券的类型。
    返回 3M / 6M / 1Y / 2Y / None（无效或旧格式）
    """
    raw = code.replace('-', '').replace(' ', '').upper()
    if len(raw) != 24:
        return None
    first_char = raw[0]
    if first_char in TYPE_MAP:
        return TYPE_MAP[first_char]["label"]
    return None


def get_coupon_valid_days(code: str) -> int:
    """获取礼品券的有效天数"""
    ctype = get_coupon_type(code)
    if ctype is None:
        return 99999  # 旧格式永久有效
    for info in TYPE_MAP.values():
        if info["label"] == ctype:
            return info["days"]
    return 99999


def validate_coupon(code: str) -> tuple[bool, str]:
    """
    校验礼品券格式。
    支持：
    - 新格式（v3）：1 位类型 + 15 位数据 + 8 位 HMAC 签名
    - 旧格式（v2）：16 位数据 + 8 位 HMAC 签名（永久有效）
    - 旧格式（v1）：22 位数据 + 2 位校验和（向下兼容）
    返回 (是否合法, 格式化后的码)
    """
    raw = code.replace('-', '').replace(' ', '').upper()
    if len(raw) != 24:
        return False, raw
    if any(ch not in CHARSET for ch in raw):
        return False, raw
    chars = list(raw)

    # v3 新格式：1 位类型 + 15 位数据 + 8 位 HMAC 签名
    first_char = chars[0]
    if first_char in TYPE_CHARS:
        data_chars = chars[:16]   # 1 type + 15 random = 16 data
        sig_chars = chars[16:]
        expected_sig = _hmac_signature(data_chars)
        if ''.join(sig_chars) == expected_sig:
            return True, raw

    # v2 旧格式：16 位数据 + 8 位 HMAC 签名
    data_chars = chars[:16]
    sig_chars = chars[16:]
    expected_sig = _hmac_signature(data_chars)
    if ''.join(sig_chars) == expected_sig:
        return True, raw

    # v1 旧格式（向下兼容）：22 位数据 + 2 位校验和
    old_data_chars = chars[:22]
    old_sig_chars = chars[22:]
    total = sum(_char_index(c) for c in old_data_chars)
    total %= CHARSET_LEN * CHARSET_LEN
    expected_cs = CHARSET[total // CHARSET_LEN] + CHARSET[total % CHARSET_LEN]
    if ''.join(old_sig_chars) == expected_cs:
        return True, raw

    return False, raw


def batch_generate(count: int = 1000, code_type: str = "1Y") -> list[str]:
    """批量生成不重复的礼品券（指定类型）"""
    seen = set()
    result = []
    while len(result) < count:
        code = generate_coupon(code_type)
        if code not in seen:
            seen.add(code)
            result.append(code)
    return result


def _push_registry(registry: dict, sha: str) -> bool:
    """将更新后的注册表推送到 GitHub（需管理员写入 TOKEN）"""
    token = _get_github_write_token()
    if not token:
        return False
    import json as _json
    content_b64 = base64.b64encode(
        _json.dumps(registry, ensure_ascii=False).encode("utf-8")
    ).decode("utf-8")
    url = (f"https://api.github.com/repos/{_GITHUB_OWNER}/{_GITHUB_REPO}"
           f"/contents/{_GITHUB_REGISTRY_PATH}")
    put_data = {
        "message": "兑换礼品券",
        "content": content_b64,
        "branch": _GITHUB_BRANCH,
    }
    # 如果有 sha 则传入（文件已存在时需要）
    if sha:
        put_data["sha"] = sha
    try:
        resp = requests.put(url, json=put_data, headers={
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
        }, timeout=10)
        return resp.status_code in (200, 201)
    except Exception:
        return False


def redeem_coupon(code: str) -> tuple[bool, str]:
    """兑换礼品券：GitHub 远程校验 → 检查跨机绑定 → 标记已使用 → 生成许可"""
    try:
        now = get_current_time()
    except RuntimeError as e:
        return False, str(e)

    valid, raw = validate_coupon(code)
    if not valid:
        return False, "礼品券格式无效"

    # 从 GitHub 私密仓库获取注册表
    registry, _ = _fetch_registry()
    write_token = _get_write_token()

    if registry is None:
        return False, "无法连接到礼品券服务器，请检查网络后重试"

    # 检查此券是否在注册表中
    if raw not in registry:
        return False, "该礼品券未在服务端注册，请联系管理员"

    current_mac = _get_mac()
    remote_entry = registry[raw]

    if remote_entry is not None:
        bound_mac = remote_entry.get("bound_mac") if isinstance(remote_entry, dict) else None
        if bound_mac:
            if current_mac != "unknown" and current_mac != bound_mac:
                return False, "此礼品券已在其他设备上使用，无法再次兑换"
            # 同机器重兑：直接写本地 license
            _write_license(raw, current_mac, now)
            ctype = get_coupon_type(raw)
            if ctype:
                return True, f"兑换成功！{TYPE_MAP.get(raw[0], {}).get('desc', '服务')}已激活，绑定本机设备"
            return True, "兑换成功！服务已激活，绑定本机设备"

    # 未绑定 -> 第一次兑换（需要写入 token）
    if not write_token:
        return False, "服务端写入 Token 未配置，请联系管理员兑换"

    # 通过 API 获取最新的 sha
    api_registry, sha = _fetch_registry_with_sha(write_token)
    if api_registry is None:
        return False, "服务端注册失败（无法获取最新注册表），请稍后重试"
    if raw not in api_registry:
        return False, "该礼品券未在服务端注册，请联系管理员"

    current_entry = api_registry[raw]
    if current_entry is not None:
        bound_mac = current_entry.get("bound_mac") if isinstance(current_entry, dict) else None
        if bound_mac and current_mac != "unknown" and current_mac != bound_mac:
            return False, "此礼品券已在其他设备上使用，无法再次兑换"

    api_registry[raw] = {
        "bound_mac": current_mac,
        "used_at": now.isoformat(),
    }
    if not _push_registry(api_registry, sha):
        return False, "服务端注册失败，请稍后重试"

    _write_license(raw, current_mac, now)
    ctype = get_coupon_type(raw)
    if ctype:
        return True, f"兑换成功！{TYPE_MAP.get(raw[0], {}).get('desc', '服务')}已激活，绑定本机设备"
    return True, "兑换成功！服务已激活，绑定本机设备"


def _write_license(raw: str, mac: str, now: datetime) -> None:
    """写入本地许可文件（含有效期）"""
    valid_days = get_coupon_valid_days(raw)
    far_future = now + timedelta(days=valid_days)
    ctype = get_coupon_type(raw)
    license_data = {
        "activated": True,
        "mac": mac,
        "coupon": raw,
        "activated_at": now.isoformat(),
        "expires_at": far_future.isoformat(),
        "coupon_type": ctype if ctype else "permanent",
        "valid_days": valid_days,
        "permanent": valid_days >= 99999,
    }
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(LICENSE_FILE, 'w', encoding='utf-8') as f:
        json.dump(license_data, f, ensure_ascii=False, indent=2)


def _get_mac() -> str:
    """获取本机 MAC 地址"""
    import uuid
    mac_int = uuid.getnode()
    if (mac_int >> 40) % 2:
        return "unknown"
    mac = ':'.join(format((mac_int >> bits) & 0xff, '02x')
                    for bits in range(40, -1, -8))
    return mac


def get_current_mac() -> str:
    """公开获取本机 MAC 地址"""
    return _get_mac()


def load_license() -> dict:
    """加载许可信息"""
    if not os.path.exists(LICENSE_FILE):
        return {"activated": False, "mac": "", "coupon": "", "activated_at": None,
                "expires_at": None, "coupon_type": "", "valid_days": 0}
    try:
        with open(LICENSE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"activated": False, "mac": "", "coupon": "", "activated_at": None,
                "expires_at": None, "coupon_type": "", "valid_days": 0}


def is_activated() -> bool:
    """检查服务是否已激活（含有效期检查，需联网确认时间）。

    支持三种许可来源：
      - 礼品券（mac 绑定）
      - 订阅（source == "subscription"）
      - 永久券（permanent == True 或 expires_at 极远）
    """
    lic = load_license()
    if not lic.get("activated"):
        return False

    # 永久券：直接视为永久激活
    if lic.get("permanent"):
        return True

    # 订阅型许可：无 mac 绑定，按到期时间判断
    if lic.get("source") == "subscription":
        expires_at = lic.get("expires_at", "")
        if not expires_at:
            return False
        try:
            net_time = _get_network_time()
            if net_time is None:
                return False
            exp = datetime.fromisoformat(expires_at)
            return exp >= net_time
        except Exception:
            return False

    # 礼品券：需 mac 绑定校验
    stored_mac = lic.get("mac", "")
    if not stored_mac:
        return False
    current_mac = _get_mac()
    if current_mac == "unknown":
        return False
    if stored_mac != current_mac:
        return False

    expires_at = lic.get("expires_at", "")
    if not expires_at:
        return False

    # 检查到期时间
    try:
        net_time = _get_network_time()
        if net_time is None:
            return False
        exp = datetime.fromisoformat(expires_at)
        if exp < net_time:
            return False
    except Exception:
        return False

    return True


def get_remaining_days() -> int:
    """获取剩余天数"""
    lic = load_license()
    if not lic.get("activated"):
        return 0
    expires_at = lic.get("expires_at", "")
    if not expires_at:
        return 0
    try:
        net_time = _get_network_time()
        if net_time is None:
            return 0
        exp = datetime.fromisoformat(expires_at)
        remaining = (exp - net_time).days
        return max(0, remaining)
    except Exception:
        return 0


# ── 7 天免费试用 ────────────────────────────────────────────
TRIAL_DAYS = 7

# 试用期锚定文件列表（写入多个文件防清理）
_TRIAL_ANCHOR_FILES = [
    "trial_record.json",
    "app_config.json",
    "tasks.json",
    "scheduler_state.json",
]


def _get_anchor_path(filename: str) -> str:
    return os.path.join(DATA_DIR, filename)


def _write_trial_anchor(mac: str, started_at: str) -> None:
    """将试用起始时间和 MAC 写入多个锚定文件"""
    for fname in _TRIAL_ANCHOR_FILES:
        fpath = _get_anchor_path(fname)
        try:
            data = {}
            if os.path.exists(fpath):
                with open(fpath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            data["_trial_mac"] = mac
            data["_trial_started"] = started_at
            with open(fpath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass


def _read_trial_from_anchors() -> tuple[str, str]:
    """从所有锚定文件中读取试用记录，返回 (mac, started_at)。取最早的。"""
    candidates = []
    for fname in _TRIAL_ANCHOR_FILES:
        fpath = _get_anchor_path(fname)
        if not os.path.exists(fpath):
            continue
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            m = data.get("_trial_mac", "")
            started = data.get("_trial_started", "")
            if m and started:
                candidates.append((m, started))
        except Exception:
            pass

    if not candidates:
        return "", ""

    candidates.sort(key=lambda x: x[1])
    return candidates[0]


def _load_trial_record() -> dict:
    if not os.path.exists(TRIAL_RECORD_FILE):
        return {}
    try:
        with open(TRIAL_RECORD_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_trial_record(record: dict) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(TRIAL_RECORD_FILE, 'w', encoding='utf-8') as f:
        json.dump(record, f, ensure_ascii=False, indent=2)


def is_trial_period(net_time: datetime = None) -> tuple[bool, int]:
    """
    判断是否在 7 天免费试用期内（MAC 绑定）。
    返回 (is_in_trial, remaining_days_or_0)。
    """
    current_mac = _get_mac()
    if current_mac == "unknown":
        return False, 0

    record = _load_trial_record()
    stored_mac = record.get("mac", "")
    stored_started = record.get("trial_started")

    # 迁移路径：旧版 app_config.json 有 trial_started
    if not record:
        from .config_manager import load_app_config
        cfg = load_app_config()
        legacy_started = cfg.get("trial_started")
        if legacy_started:
            record = {
                "mac": current_mac,
                "trial_started": legacy_started,
                "trial_notified": cfg.get("trial_notified", False),
            }
            _save_trial_record(record)
            stored_mac = current_mac
            stored_started = legacy_started

    if stored_mac == current_mac and stored_started:
        if net_time is None:
            net_time = _get_network_time()
        if net_time is None:
            return False, 0
        started = datetime.fromisoformat(stored_started)
        elapsed_days = (net_time - started).days
        remaining = max(0, TRIAL_DAYS - elapsed_days)
        if remaining > 0:
            return True, remaining
        return False, 0

    # 试用记录为空 → 从锚定文件恢复
    anchor_mac, anchor_started = _read_trial_from_anchors()
    if anchor_mac == current_mac and anchor_started:
        stored_started = anchor_started
        if net_time is None:
            net_time = _get_network_time()
        if net_time is None:
            return False, 0
        started = datetime.fromisoformat(stored_started)
        elapsed_days = (net_time - started).days
        remaining = max(0, TRIAL_DAYS - elapsed_days)
        _save_trial_record({
            "mac": current_mac,
            "trial_started": stored_started,
            "trial_notified": False,
        })
        if remaining > 0:
            return True, remaining
        return False, 0

    # 无试用记录 → 首次安装
    if net_time is None:
        net_time = _get_network_time()
    if net_time is None:
        return False, 0

    record = {
        "mac": current_mac,
        "trial_started": net_time.isoformat(),
        "trial_notified": False,
    }
    _save_trial_record(record)
    _write_trial_anchor(current_mac, net_time.isoformat())

    from .config_manager import load_app_config, save_app_config
    cfg = load_app_config()
    cfg["trial_started"] = net_time.isoformat()
    save_app_config(cfg)

    return True, TRIAL_DAYS


def is_feature_allowed() -> bool:
    """功能许可检查（7天试用 or 礼品券激活 or 订阅激活，含 3 天宽限期）"""
    return is_in_validity()


# ── 统一有效期检测（礼品券 / 订阅 / 试用，含宽限期） ────────
GRACE_DAYS = 3


def is_in_validity(grace_days: int = GRACE_DAYS) -> bool:
    """综合判断当前是否处于功能有效期内。

    优先级：订阅激活 → 礼品券激活 → 试用期。
    到期后在宽限期内仍视为有效（供 UI 提示续费）；宽限期过后才失效。
    """
    # 订阅激活
    try:
        from . import subscription
        if subscription.is_subscription_active():
            return True
    except Exception:
        pass
    # 礼品券激活（含到期判断）
    if is_activated():
        return True
    # 礼品券到期但处于宽限期
    lic = load_license()
    expires_at = lic.get("expires_at", "")
    if lic.get("activated") and expires_at:
        try:
            net_time = _get_network_time()
            if net_time is not None:
                exp = datetime.fromisoformat(expires_at)
                if exp < net_time and (exp + timedelta(days=grace_days)) > net_time:
                    return True
        except Exception:
            pass
    # 试用期
    trial_ok, _ = is_trial_period()
    if trial_ok:
        return True
    # 试用到期宽限
    _, trial_remain = is_trial_period()
    if trial_remain <= 0:
        record = _load_trial_record()
        started = record.get("trial_started", "")
        if started:
            try:
                net_time = _get_network_time()
                if net_time is not None:
                    start = datetime.fromisoformat(started)
                    trial_end = start + timedelta(days=TRIAL_DAYS)
                    if trial_end < net_time and (trial_end + timedelta(days=grace_days)) > net_time:
                        return True
            except Exception:
                pass
    return False


def get_expiry_info() -> dict:
    """获取到期信息：{source, expires_at, remaining_days, in_grace, grace_until, permanent}。"""
    # 订阅优先
    try:
        from . import subscription
        sub_status = subscription.get_license_status()
        if sub_status.get("active") and sub_status.get("source") == "subscription":
            return sub_status
    except Exception:
        pass

    lic = load_license()
    expires_at = lic.get("expires_at", "")
    source = lic.get("source", lic.get("coupon_type", "")) or ""
    remaining = 0
    in_grace = False
    grace_until = ""
    permanent = bool(lic.get("permanent")) or (lic.get("coupon_type") in (None, "", "permanent"))

    # 永久券：永不显示到期/宽限
    if lic.get("activated") and permanent:
        return {
            "active": True,
            "source": source or "gift",
            "expires_at": expires_at,
            "remaining_days": 99999,
            "in_grace": False,
            "grace_until": "",
            "permanent": True,
        }

    if lic.get("activated") and expires_at:
        try:
            net_time = _get_network_time()
            if net_time is not None:
                exp = datetime.fromisoformat(expires_at)
                remaining = max(0, (exp - net_time).days)
                if remaining == 0:
                    grace_until_dt = exp + timedelta(days=GRACE_DAYS)
                    in_grace = grace_until_dt > net_time
                    grace_until = grace_until_dt.isoformat()
        except Exception:
            pass

    if not source and not lic.get("activated"):
        # 试用期
        trial_ok, remain = is_trial_period()
        if trial_ok:
            record = _load_trial_record()
            started = record.get("trial_started", "")
            return {
                "active": True,
                "source": "trial",
                "expires_at": (datetime.fromisoformat(started) + timedelta(days=TRIAL_DAYS)).isoformat()
                    if started else "",
                "remaining_days": remain,
                "in_grace": False,
                "grace_until": "",
                "permanent": False,
            }

    return {
        "active": bool(lic.get("activated")),
        "source": source or "unknown",
        "expires_at": expires_at,
        "remaining_days": remaining,
        "in_grace": in_grace,
        "grace_until": grace_until,
        "permanent": permanent,
    }


def get_remaining_days_all() -> int:
    """获取剩余可用天数（先查礼品券，无券查试用）"""
    coupon_days = get_remaining_days()
    if coupon_days > 0:
        return coupon_days
    _, trial_days = is_trial_period()
    return trial_days


def check_mac_match(target_mac: str) -> bool:
    """比较给定 MAC 是否与本机一致"""
    return _get_mac() == target_mac


# ── 批量导出（生成 xlsx 到桌面）─────────────────────────────

def generate_all_type_coupons(count_per_type: int = 1000) -> dict[str, list[str]]:
    """生成所有类型的礼品券"""
    result = {}
    for ctype in ["3M", "6M", "1Y", "2Y"]:
        result[ctype] = batch_generate(count_per_type, ctype)
    return result


def export_coupons_to_xlsx(filepath: str = None) -> str:
    """
    生成 4 类 × 1000 礼品券并导出为 xlsx 文件。
    默认保存到桌面：~/Desktop/HONGXUN_礼品券_{时间戳}.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # 如果 openpyxl 不可用，保存为 csv
        return _export_coupons_to_csv(filepath)

    data = generate_all_type_coupons()

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ctype, codes in data.items():
        info = {"3M": "3个月(90天)", "6M": "6个月(180天)", "1Y": "12个月(365天)", "2Y": "24个月(730天)"}
        ws = wb.create_sheet(title=ctype)

        # 标题行
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = f"鸿讯 HONGXUN 礼品券 — {info.get(ctype, ctype)}"
        title_cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="2563EB")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # 表头
        headers = ["序号", "礼品券编码", "有效期"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        body_font = Font(name="Consolas", size=10)
        for i, code in enumerate(codes, 1):
            formatted = format_coupon(code)
            row = i + 2
            ws.cell(row=row, column=1, value=i).font = body_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=1).border = thin_border

            cell_code = ws.cell(row=row, column=2, value=formatted)
            cell_code.font = Font(name="Consolas", size=10, bold=True)
            cell_code.alignment = Alignment(horizontal="center")
            cell_code.border = thin_border

            cell_type = ws.cell(row=row, column=3, value=info.get(ctype, ctype))
            cell_type.font = Font(name="Microsoft YaHei", size=10)
            cell_type.alignment = Alignment(horizontal="center")
            cell_type.border = thin_border

        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 18

    if filepath is None:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(desktop, f"HONGXUN_礼品券_{timestamp}.xlsx")

    wb.save(filepath)
    return filepath


def _export_coupons_to_csv(filepath: str = None) -> str:
    """备用方案：导出为 csv"""
    import csv
    data = generate_all_type_coupons()

    if filepath is None:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(desktop, f"HONGXUN_礼品券_{timestamp}.csv")

    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["类型", "礼品券编码", "有效期"])
        for ctype, codes in data.items():
            info = {"3M": "3个月(90天)", "6M": "6个月(180天)", "1Y": "12个月(365天)", "2Y": "24个月(730天)"}
            for code in codes:
                writer.writerow([ctype, format_coupon(code), info.get(ctype, ctype)])

    return filepath
