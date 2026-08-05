"""
鸿讯 HONGXUN · 期刊批量导入弹窗
版本 1.0.0

支持从多种来源批量导入期刊到已选列表：
  1. 手动粘贴：每行 / 分号 / 逗号分隔的期刊名
  2. RIS 文件（EndNote / Zotero）
  3. BibTeX 文件
  4. Excel / CSV 文件（取"期刊"或"Journal"列）

导入时自动在期刊数据库模糊匹配，展示命中结果，用户勾选后回填。
"""

import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from gui.theme import COLORS, FONT_BODY, FONT_BODY_BOLD, FONT_CAPTION, FONT_HEADING
from gui.widgets import ModernButton, RoundedCard, ModernScrollbar


class JournalImportDialog(tk.Toplevel):
    """期刊批量导入弹窗。"""

    def __init__(self, master, store, on_import=None):
        super().__init__(master)
        self.title("批量导入期刊")
        self.geometry("640x560")
        self.minsize(560, 440)
        self.configure(bg=COLORS["bg_page"])
        self.transient(master)
        self.grab_set()

        self.store = store
        self.on_import = on_import  # on_import(list[dict] 匹配到的期刊)

        # 匹配结果：[(journal dict or None, raw_name)]
        self._results = []

        self._build_ui()
        self._place_center()

    def _place_center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (sw - w) // 2, max((sh - h) // 2 - 40, 40)
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        # ── 顶部：来源选择 ──
        header = tk.Frame(self, bg=COLORS["bg_page"])
        header.pack(fill=tk.X, padx=16, pady=(16, 8))
        tk.Label(header, text="批量导入期刊", font=FONT_HEADING,
                 fg=COLORS["text_title"], bg=COLORS["bg_page"]).pack(side=tk.LEFT)

        self._mode_var = tk.StringVar(value="paste")
        mode_frame = tk.Frame(self, bg=COLORS["bg_page"])
        mode_frame.pack(fill=tk.X, padx=16)
        for key, label in (("paste", "✏️ 手动粘贴"), ("file", "📄 文件导入")):
            tk.Radiobutton(mode_frame, text=label, variable=self._mode_var,
                           value=key, command=self._on_mode_change,
                           bg=COLORS["bg_page"], fg=COLORS["text_body"],
                           selectcolor=COLORS["bg_page"],
                           activebackground=COLORS["bg_page"]).pack(side=tk.LEFT, padx=(0, 16))

        # ── 输入区（粘贴 or 文件） ──
        self._input_card = RoundedCard(self, radius=12, pad=0, shadow=True,
                                       bg_color=COLORS["bg_card"], fit_content=True)
        self._input_card.pack(fill=tk.X, padx=16, pady=(8, 10))
        self._build_paste_area(self._input_card.content)
        self._file_area = tk.Frame(self._input_card.content, bg=COLORS["bg_card"])
        self._build_file_area(self._file_area)

        # ── 结果列表（滚动） ──
        result_card = RoundedCard(self, radius=12, pad=0, shadow=True,
                                  bg_color=COLORS["bg_card"])
        result_card.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 10))
        tk.Label(result_card.content, text="匹配结果", font=FONT_HEADING,
                 fg=COLORS["text_title"], bg=COLORS["bg_card"]).pack(
            anchor=tk.W, padx=14, pady=(12, 4))
        tk.Frame(result_card.content, bg=COLORS["border_light"], height=1).pack(
            fill=tk.X, padx=14, pady=(0, 6))

        self._list_container = tk.Frame(result_card.content, bg=COLORS["bg_card"])
        self._list_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(4, 12))
        self._canvas = tk.Canvas(self._list_container, borderwidth=0,
                                 highlightthickness=0, bg=COLORS["bg_card"])
        vsb = ModernScrollbar(self._list_container, width=8,
                              command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._results_frame = tk.Frame(self._canvas, bg=COLORS["bg_card"])
        self._win = self._canvas.create_window((0, 0), window=self._results_frame,
                                               anchor=tk.NW)
        self._results_frame.bind(
            "<Configure>", lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._canvas.bind(
            "<Configure>", lambda e: self._canvas.itemconfig(self._win, width=e.width))
        self._canvas.bind("<MouseWheel>",
                          lambda e: self._canvas.yview_scroll(int(-e.delta / 120), "units"))

        # ── 底部按钮 ──
        bottom = tk.Frame(self, bg=COLORS["bg_page"])
        bottom.pack(fill=tk.X, padx=16, pady=(0, 14))
        self._summary_label = tk.Label(bottom, text="", font=FONT_CAPTION,
                                       fg=COLORS["text_hint"], bg=COLORS["bg_page"])
        self._summary_label.pack(side=tk.LEFT)
        ModernButton(bottom, text="取消", variant="secondary",
                     command=self.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        ModernButton(bottom, text="导入选中", variant="primary",
                     command=self._on_confirm).pack(side=tk.RIGHT)

    def _build_paste_area(self, parent):
        self._text = tk.Text(parent, height=7, bd=0, relief="flat",
                             highlightthickness=1,
                             highlightbackground=COLORS["input_border"],
                             highlightcolor=COLORS["primary"],
                             bg=COLORS["bg_input"], fg=COLORS["text_body"],
                             font=FONT_BODY, wrap=tk.WORD)
        self._text.pack(fill=tk.X, padx=14, pady=(10, 4))
        hint = tk.Frame(parent, bg=COLORS["bg_card"])
        hint.pack(fill=tk.X, padx=14, pady=(0, 10))
        tk.Label(hint, text="每行一个期刊名，或分号/逗号分隔", font=FONT_CAPTION,
                 fg=COLORS["text_hint"], bg=COLORS["bg_card"]).pack(side=tk.LEFT)
        ModernButton(hint, text="开始匹配", variant="primary", height=28, pad_x=14,
                     command=self._match_paste).pack(side=tk.RIGHT)

    def _build_file_area(self, parent):
        tk.Label(parent, text="选择文件（支持 .ris / .bib / .xlsx / .csv）", font=FONT_CAPTION,
                 fg=COLORS["text_hint"], bg=COLORS["bg_card"]).pack(
            anchor=tk.W, padx=14, pady=(10, 6))
        row = tk.Frame(parent, bg=COLORS["bg_card"])
        row.pack(fill=tk.X, padx=14, pady=(0, 10))
        self._file_label = tk.Label(row, text="未选择文件", font=FONT_BODY,
                                    fg=COLORS["text_secondary"], bg=COLORS["bg_card"],
                                    anchor=tk.W)
        self._file_label.pack(side=tk.LEFT, expand=True, fill=tk.X)
        ModernButton(row, text="选择文件…", variant="secondary", height=28, pad_x=14,
                     command=self._choose_file).pack(side=tk.RIGHT)
        ModernButton(row, text="开始匹配", variant="primary", height=28, pad_x=14,
                     command=self._match_file).pack(side=tk.RIGHT, padx=(8, 0))

    def _on_mode_change(self):
        if self._mode_var.get() == "paste":
            self._file_area.pack_forget()
            self._text.master.pack(fill=tk.X, padx=14, pady=(10, 4))
            for w in self._input_card.content.winfo_children():
                if w is not self._text.master and w is not self._file_area:
                    w.pack(fill=tk.X, padx=14, pady=(0, 10))
        else:
            for w in self._input_card.content.winfo_children():
                if w is not self._file_area:
                    w.pack_forget()
            self._file_area.pack(fill=tk.X, padx=0, pady=0)

    # ── 匹配 ─────────────────────────────────────────────
    def _match_paste(self):
        text = self._text.get("1.0", tk.END).strip() if self._text else ""
        if not text:
            messagebox.showinfo("提示", "请先粘贴期刊名称", parent=self)
            return
        names = self._parse_names(text)
        self._run_match(names)

    def _match_file(self):
        path = self._file_path if hasattr(self, "_file_path") else ""
        if not path or not os.path.exists(path):
            messagebox.showinfo("提示", "请先选择文件", parent=self)
            return
        names = self._parse_file(path)
        if not names:
            messagebox.showinfo("提示", "文件中未识别到期刊名称", parent=self)
            return
        self._run_match(names)

    @staticmethod
    def _parse_names(text: str) -> list[str]:
        """从文本中提取期刊名列表（兼容行、分号、逗号、顿号分隔）。"""
        names = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            for part in re.split(r"[;；,，、]", line):
                part = part.strip()
                if part:
                    names.append(part)
        # 去重保序
        seen = set()
        out = []
        for n in names:
            key = n.lower()
            if key not in seen:
                seen.add(key)
                out.append(n)
        return out

    def _parse_file(self, path: str) -> list[str]:
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".ris",):
                with open(path, encoding="utf-8", errors="ignore") as f:
                    content = f.read()
                # RIS：JF / JO / T2 为期刊字段
                matches = re.findall(r"^(JF|JO|T2)\s{2}-\s*(.+)$", content, re.MULTILINE)
                names = [m[1].strip() for m in matches]
            elif ext in (".bib",):
                with open(path, encoding="utf-8", errors="ignore") as f:
                    content = f.read()
                matches = re.findall(r"journal\s*=\s*[{\"](.+?)[}\"]", content,
                                     re.IGNORECASE)
                names = [m.strip() for m in matches]
            elif ext in (".csv",):
                import csv
                names = []
                with open(path, encoding="utf-8-sig", errors="ignore") as f:
                    for row in csv.DictReader(f):
                        for key in ("期刊", "Journal", "期刊名", "journal"):
                            if key in row and row[key].strip():
                                names.append(row[key].strip())
                                break
            elif ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True)
                    ws = wb.active
                    names = []
                    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    col_idx = None
                    for i, h in enumerate(header):
                        if h.lower() in ("期刊", "journal", "期刊名", "期刊名称"):
                            col_idx = i
                            break
                    for row in ws.iter_rows(min_row=2):
                        if col_idx is not None and row[col_idx].value:
                            names.append(str(row[col_idx].value).strip())
                    wb.close()
                except ImportError:
                    messagebox.showwarning("缺少依赖", "读取 Excel 需要安装 openpyxl，请先 pip install openpyxl", parent=self)
                    return []
            else:
                return []
        except Exception as e:
            messagebox.showerror("读取失败", f"无法读取文件：{e}", parent=self)
            return []
        return self._parse_names("\n".join(names))

    def _run_match(self, names: list[str]):
        """批量匹配期刊名到数据库。"""
        self._results = []
        for name in names:
            j = None
            try:
                j = self.store.get_by_name(name, fuzzy=True)
            except Exception:
                j = None
            self._results.append((j, name))
        self._render_results()

    def _render_results(self):
        for w in self._results_frame.winfo_children():
            w.destroy()

        matched = sum(1 for j, _ in self._results if j)
        self._summary_label.configure(
            text=f"共 {len(self._results)} 条，匹配 {matched} 条")

        self._checks = []
        for j, raw in self._results:
            item = tk.Frame(self._results_frame, bg=COLORS["bg_card"])
            item.pack(fill=tk.X, pady=2)

            var = tk.BooleanVar(value=j is not None)
            self._checks.append(var)
            cb = tk.Checkbutton(item, variable=var, bg=COLORS["bg_card"],
                                activebackground=COLORS["bg_card"],
                                selectcolor=COLORS["bg_page"])
            cb.pack(side=tk.LEFT)

            if j:
                tk.Label(item, text="✅", font=FONT_BODY,
                         bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=(0, 4))
                tk.Label(item, text=j["full_name"], font=FONT_BODY_BOLD,
                         fg=COLORS["text_body"], bg=COLORS["bg_card"],
                         anchor=tk.W).pack(side=tk.LEFT)
                d = j.get("cas_division_2024") or 0
                tag = f"  {d}区" if 1 <= d <= 4 else ""
                if j.get("is_top"):
                    tag += " TOP"
                tag += f"  IF {j.get('impact_factor_2025') or 0:.1f}"
                tk.Label(item, text=tag, font=FONT_CAPTION,
                         fg=COLORS["text_hint"], bg=COLORS["bg_card"]).pack(
                    side=tk.LEFT, padx=(8, 0))
            else:
                tk.Label(item, text="⚠️", font=FONT_BODY,
                         bg=COLORS["bg_card"]).pack(side=tk.LEFT, padx=(0, 4))
                tk.Label(item, text=raw, font=FONT_BODY,
                         fg=COLORS["text_hint"], bg=COLORS["bg_card"],
                         anchor=tk.W).pack(side=tk.LEFT)
                tk.Label(item, text="未匹配", font=FONT_CAPTION,
                         fg=COLORS["warning"], bg=COLORS["bg_card"]).pack(
                    side=tk.LEFT, padx=(8, 0))

    def _choose_file(self):
        path = filedialog.askopenfilename(
            parent=self,
            filetypes=[("文献文件", "*.ris *.bib *.xlsx *.xls *.csv"), ("所有文件", "*.*")])
        if path:
            self._file_path = path
            self._file_label.configure(text=os.path.basename(path))

    def _on_confirm(self):
        selected = []
        for var, (j, raw) in zip(self._checks, self._results):
            if var.get() and j:
                selected.append(j)
        if not selected:
            messagebox.showinfo("提示", "请至少勾选一个匹配到的期刊", parent=self)
            return
        if self.on_import:
            try:
                self.on_import(selected)
            except Exception:
                pass
        self.destroy()
